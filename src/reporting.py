"""Excel MIS reporting and dashboard generation."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database import AccessRequest, Exception, PROJECT_ROOT, User
from rules_engine import get_exception_summary

REPORTS_DIR = PROJECT_ROOT / "reports"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, min_width: int = 12) -> None:
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(length + 2, min_width)


def _request_metrics(session: Session) -> dict:
    requests = session.query(AccessRequest).all()
    df = pd.DataFrame(
        [
            {
                "request_id": r.request_id,
                "user_id": r.user_id,
                "requested_system": r.requested_system,
                "request_date": r.request_date,
                "approval_status": r.approval_status,
                "closure_status": r.closure_status,
            }
            for r in requests
        ]
    )

    if df.empty:
        return {
            "total": 0,
            "approved": 0,
            "rejected": 0,
            "pending": 0,
            "avg_tat_days": 0.0,
            "by_department": pd.DataFrame(),
            "monthly_trends": pd.DataFrame(),
            "requests_df": df,
        }

    df["request_date"] = pd.to_datetime(df["request_date"])
    df["month"] = df["request_date"].dt.to_period("M").astype(str)

    users = pd.DataFrame(
        [{"user_id": u.user_id, "department": u.department} for u in session.query(User).all()]
    )
    merged = df.merge(users, on="user_id", how="left")

    closed = df[df["approval_status"].isin(["Approved", "Rejected"])].copy()
    if not closed.empty:
        closed["tat_days"] = (pd.Timestamp(date.today()) - closed["request_date"]).dt.days.clip(lower=1)
        avg_tat = round(closed["tat_days"].mean(), 1)
    else:
        avg_tat = 0.0

    by_dept = (
        merged.groupby("department")
        .size()
        .reset_index(name="request_count")
        .sort_values("request_count", ascending=False)
    )

    monthly = (
        df.groupby("month")
        .size()
        .reset_index(name="request_count")
        .sort_values("month")
    )

    return {
        "total": len(df),
        "approved": int((df["approval_status"] == "Approved").sum()),
        "rejected": int((df["approval_status"] == "Rejected").sum()),
        "pending": int((df["approval_status"] == "Pending").sum()),
        "avg_tat_days": avg_tat,
        "by_department": by_dept,
        "monthly_trends": monthly,
        "requests_df": merged,
    }


def export_exception_report(session: Session, output_path: Path | None = None) -> Path:
    output_path = output_path or REPORTS_DIR / "exception_report.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    exceptions = session.query(Exception).order_by(
        Exception.severity, Exception.user_id
    ).all()

    df = pd.DataFrame(
        [
            {
                "User ID": e.user_id,
                "User Name": e.user_name,
                "Issue Type": e.issue_type,
                "Severity": e.severity,
                "Description": e.description,
            }
            for e in exceptions
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Exceptions", index=False)
        ws = writer.sheets["Exceptions"]
        if not df.empty:
            _style_header_row(ws, 1, len(df.columns))
            _auto_width(ws)

    return output_path


def export_audit_report(session: Session, output_path: Path | None = None) -> Path:
    from audit import audit_logs_to_dataframe

    output_path = output_path or REPORTS_DIR / "audit_report.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = audit_logs_to_dataframe(session)
    df.columns = ["Event ID", "User ID", "Event Type", "Timestamp", "Description"]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Audit Trail", index=False)
        ws = writer.sheets["Audit Trail"]
        if not df.empty:
            _style_header_row(ws, 1, len(df.columns))
            _auto_width(ws)

    return output_path


def export_management_dashboard(session: Session, output_path: Path | None = None) -> Path:
    output_path = output_path or REPORTS_DIR / "management_dashboard.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    metrics = _request_metrics(session)
    exc_summary = get_exception_summary(session)

    wb = Workbook()

    # --- Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "System Access Governance — MIS Dashboard"
    ws_summary["A1"].font = TITLE_FONT
    ws_summary["A2"] = f"Report Date: {date.today().isoformat()}"

    summary_rows = [
        ("Metric", "Value"),
        ("Total Requests", metrics["total"]),
        ("Approved Requests", metrics["approved"]),
        ("Rejected Requests", metrics["rejected"]),
        ("Pending Requests", metrics["pending"]),
        ("Average Turnaround Time (days)", metrics["avg_tat_days"]),
        ("Total Exceptions", exc_summary["total"]),
        ("High Severity Exceptions", exc_summary["by_severity"].get("High", 0)),
        ("Medium Severity Exceptions", exc_summary["by_severity"].get("Medium", 0)),
        ("Low Severity Exceptions", exc_summary["by_severity"].get("Low", 0)),
    ]

    for i, (label, value) in enumerate(summary_rows, start=4):
        ws_summary.cell(row=i, column=1, value=label)
        ws_summary.cell(row=i, column=2, value=value)
        if i == 4:
            _style_header_row(ws_summary, i, 2)

    # Request status pie chart
    if metrics["total"] > 0:
        ws_summary["D4"] = "Status"
        ws_summary["E4"] = "Count"
        status_data = [
            ("Approved", metrics["approved"]),
            ("Rejected", metrics["rejected"]),
            ("Pending", metrics["pending"]),
        ]
        for idx, (status, count) in enumerate(status_data, start=5):
            ws_summary.cell(row=idx, column=4, value=status)
            ws_summary.cell(row=idx, column=5, value=count)

        pie = PieChart()
        pie.title = "Request Status Distribution"
        pie.add_data(Reference(ws_summary, min_col=5, min_row=4, max_row=7), titles_from_data=True)
        pie.set_categories(Reference(ws_summary, min_col=4, min_row=5, max_row=7))
        pie.width = 14
        pie.height = 10
        ws_summary.add_chart(pie, "G4")

    _auto_width(ws_summary)

    # --- Operations Sheet ---
    ws_ops = wb.create_sheet("Operations")
    ws_ops["A1"] = "Top Departments by Access Requests"
    ws_ops["A1"].font = TITLE_FONT

    dept_df = metrics["by_department"]
    ws_ops.append(["Department", "Request Count"])
    _style_header_row(ws_ops, 2, 2)
    for _, row in dept_df.iterrows():
        ws_ops.append([row["department"], row["request_count"]])

    if not dept_df.empty:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Requests by Department"
        bar.y_axis.title = "Requests"
        bar.x_axis.title = "Department"
        data = Reference(ws_ops, min_col=2, min_row=2, max_row=1 + len(dept_df))
        cats = Reference(ws_ops, min_col=1, min_row=3, max_row=1 + len(dept_df))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.width = 18
        bar.height = 10
        ws_ops.add_chart(bar, "D2")

    ws_ops.append([])
    ws_ops.append(["Monthly Access Request Trends"])
    ws_ops.cell(row=ws_ops.max_row, column=1).font = TITLE_FONT

    monthly = metrics["monthly_trends"]
    trend_start = ws_ops.max_row + 1
    ws_ops.append(["Month", "Request Count"])
    _style_header_row(ws_ops, trend_start, 2)
    for _, row in monthly.iterrows():
        ws_ops.append([row["month"], row["request_count"]])

    if not monthly.empty:
        line = LineChart()
        line.title = "Monthly Access Trends"
        line.y_axis.title = "Requests"
        line.x_axis.title = "Month"
        data = Reference(ws_ops, min_col=2, min_row=trend_start, max_row=trend_start + len(monthly))
        cats = Reference(ws_ops, min_col=1, min_row=trend_start + 1, max_row=trend_start + len(monthly))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.width = 18
        line.height = 10
        ws_ops.add_chart(line, f"D{trend_start}")

    _auto_width(ws_ops)

    # --- Exceptions Sheet ---
    ws_exc = wb.create_sheet("Exceptions")
    ws_exc["A1"] = "Exception Analysis"
    ws_exc["A1"].font = TITLE_FONT

    severity_data = exc_summary["by_severity"]
    ws_exc.append(["Severity", "Count"])
    _style_header_row(ws_exc, 2, 2)
    for sev in ["High", "Medium", "Low"]:
        ws_exc.append([sev, severity_data.get(sev, 0)])

    if severity_data:
        bar_exc = BarChart()
        bar_exc.title = "Exceptions by Severity"
        bar_exc.y_axis.title = "Count"
        data = Reference(ws_exc, min_col=2, min_row=2, max_row=4)
        cats = Reference(ws_exc, min_col=1, min_row=3, max_row=4)
        bar_exc.add_data(data, titles_from_data=True)
        bar_exc.set_categories(cats)
        bar_exc.width = 14
        bar_exc.height = 10
        ws_exc.add_chart(bar_exc, "D2")

    ws_exc.append([])
    dept_exc_start = ws_exc.max_row + 1
    ws_exc.cell(row=dept_exc_start, column=1, value="Control Breaches by Department").font = TITLE_FONT
    dept_exc_start += 1
    ws_exc.append(["Department", "Breach Count"])
    _style_header_row(ws_exc, dept_exc_start, 2)

    dept_breaches = exc_summary["by_department"]
    for dept, count in sorted(dept_breaches.items(), key=lambda x: -x[1]):
        ws_exc.append([dept, count])

    if dept_breaches:
        bar_dept = BarChart()
        bar_dept.title = "Breaches by Department"
        n = len(dept_breaches)
        data = Reference(ws_exc, min_col=2, min_row=dept_exc_start, max_row=dept_exc_start + n)
        cats = Reference(ws_exc, min_col=1, min_row=dept_exc_start + 1, max_row=dept_exc_start + n)
        bar_dept.add_data(data, titles_from_data=True)
        bar_dept.set_categories(cats)
        bar_dept.width = 16
        bar_dept.height = 10
        ws_exc.add_chart(bar_dept, f"D{dept_exc_start}")

    _auto_width(ws_exc)

    # --- Trend Analysis Sheet ---
    ws_trend = wb.create_sheet("Trend Analysis")
    ws_trend["A1"] = "Governance Trend Analysis"
    ws_trend["A1"].font = TITLE_FONT

    type_data = exc_summary["by_type"]
    ws_trend.append(["Issue Type", "Count"])
    _style_header_row(ws_trend, 2, 2)
    for issue, count in sorted(type_data.items(), key=lambda x: -x[1]):
        ws_trend.append([issue, count])

    if type_data:
        bar_type = BarChart()
        bar_type.title = "Exception Types"
        n = len(type_data)
        data = Reference(ws_trend, min_col=2, min_row=2, max_row=1 + n)
        cats = Reference(ws_trend, min_col=1, min_row=3, max_row=1 + n)
        bar_type.add_data(data, titles_from_data=True)
        bar_type.set_categories(cats)
        bar_type.width = 20
        bar_type.height = 12
        ws_trend.add_chart(bar_type, "D2")

    req_df = metrics["requests_df"]
    if not req_df.empty:
        monthly_status = (
            req_df.groupby(["month", "approval_status"])
            .size()
            .unstack(fill_value=0)
            .reset_index()
        )
        ws_trend.append([])
        start = ws_trend.max_row + 1
        ws_trend.cell(row=start, column=1, value="Monthly Status Breakdown").font = TITLE_FONT
        headers = ["Month"] + list(monthly_status.columns[1:])
        ws_trend.append(headers)
        _style_header_row(ws_trend, start + 1, len(headers))
        for _, row in monthly_status.iterrows():
            ws_trend.append([row["month"]] + [row[c] for c in monthly_status.columns[1:]])

    _auto_width(ws_trend)
    wb.save(output_path)
    return output_path


def generate_matplotlib_charts(session: Session, output_dir: Path | None = None) -> list[Path]:
    """Generate supplementary PNG charts for documentation and dashboards."""
    import matplotlib.pyplot as plt

    output_dir = output_dir or REPORTS_DIR / "charts"
    output_dir.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []

    metrics = _request_metrics(session)
    exc_summary = get_exception_summary(session)

    # Request status chart
    fig, ax = plt.subplots(figsize=(8, 5))
    labels = ["Approved", "Rejected", "Pending"]
    values = [metrics["approved"], metrics["rejected"], metrics["pending"]]
    colors = ["#2E7D32", "#C62828", "#F9A825"]
    ax.bar(labels, values, color=colors)
    ax.set_title("Access Request Status Distribution")
    ax.set_ylabel("Count")
    path = output_dir / "request_status.png"
    fig.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)
    saved.append(path)

    # Exception severity chart
    if exc_summary["by_severity"]:
        fig, ax = plt.subplots(figsize=(8, 5))
        sevs = ["High", "Medium", "Low"]
        counts = [exc_summary["by_severity"].get(s, 0) for s in sevs]
        ax.bar(sevs, counts, color=["#B71C1C", "#EF6C00", "#FBC02D"])
        ax.set_title("Exceptions by Severity")
        ax.set_ylabel("Count")
        path = output_dir / "exception_severity.png"
        fig.tight_layout()
        fig.savefig(path, dpi=120)
        plt.close(fig)
        saved.append(path)

    return saved


def export_all_reports(session: Session) -> dict[str, Path]:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return {
        "exception_report": export_exception_report(session),
        "audit_report": export_audit_report(session),
        "management_dashboard": export_management_dashboard(session),
        "charts": generate_matplotlib_charts(session),
    }
